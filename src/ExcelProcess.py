import logging
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s : %(message)s')


class MSExcel:
    def __init__(self, file_name):      # 建構元
        self.__file_name = file_name
        self.__wb = None                # excel 檔案物件
        self.__all_sheet = None         # 全部的工作表
        self.__active_sheet = None      # 目前正在工作的工作表

    def open_file(self, type_flag=True):
        """

        :rtype: object
        :param type_flag:   # data_only : True : 取 cell 的 值 , False : 取 cell 的公式
        :return:
        """
        wb = openpyxl.load_workbook(self.__file_name, data_only=type_flag)
        self.__wb = wb                                          # excel 檔案物件
        self.__all_sheet = wb.get_sheet_names()                 # 全部的工作表
        self.__active_sheet = wb.get_active_sheet()             # 目前正在工作的工作表

        self.show_sheet()
        # self.set_sheet("2020Q3")    # 設定目前的工作表
        # log_txt = ("儲存格A1 = ", self.__active_sheet['A1'].column, self.__active_sheet['A1'].row, self.__active_sheet['A1'].coordinate)
        # logging.info(log_txt)

    def set_sheet(self, sheet):
        self.__active_sheet = self.__wb.get_sheet_by_name(sheet)    # 設定目前的工作表
        self.show_sheet()

    def get_max_col(self):
        return self.__active_sheet.max_column

    def get_max_row(self):
        return self.__active_sheet.max_row

    def get_row_data(self, parm_row=1):
        txt = []
        for i in range(1, self.__active_sheet.max_column + 1):
            test = str(self.__active_sheet.cell(column=i, row=parm_row).value)
            txt.append(test)
        return txt

    def get_col_data(self, parm_col=1):
        txt = []
        for cell in list(self.__active_sheet.columns)[parm_col-1]:
            txt.append(cell.value)
        return txt

    def get_cell_black(self,parm_start,parm_end):
        txt_list_array = []
        for row in self.__active_sheet[parm_start:parm_end]:
            txt_list = []
            for cell in row:
                txt_list.append(cell.value)
            txt_list_array.append(txt_list)
        return txt_list_array

    def save_new_excel(self, new_file, old_file="", new_sheet=""):
        if old_file == "":
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(old_file)

        if new_sheet != "":
            ws = wb.get_active_sheet()
            wb.create_sheet(title=new_sheet)
        wb.save(new_file)

    def remove_sheet(self, sheet_name):
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))

    def show_sheet(self):
        logging.info("所有的工作表: " + str(self.__all_sheet))
        logging.info("目前的工作表: " + str(self.__active_sheet))
        logging.info("目前的工作表: " + str(self.__active_sheet.title))
        logging.info("工作表欄數 = " + str(self.__active_sheet.max_column))
        logging.info("工作表行數 = " + str(self.__active_sheet.max_row))
        logging.info(get_column_letter(3))
        logging.info(column_index_from_string("AA"))
        logging.info("")

